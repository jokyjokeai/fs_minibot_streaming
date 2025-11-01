#!/usr/bin/env python3
"""
Import Contacts - MiniBotPanel v3

Import de contacts depuis CSV/Excel.

Usage:
    python import_contacts.py --source contacts.csv
    python import_contacts.py --source contacts.xlsx --campaign "Nouvelle campagne"

Expected CSV/Excel format:
    phone,first_name,last_name,company,email,notes
"""

import argparse
import logging
import csv
import sys
from pathlib import Path

from system.database import SessionLocal
from system.models import Contact, Campaign, Call, CampaignStatus, CallStatus
from system.campaign_manager import CampaignManager

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def read_csv_file(file_path: str) -> list:
    """
    Read contacts from CSV file.

    Expected format: phone,first_name,last_name,company,email,notes
    """
    contacts = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader, start=1):
                # Validate required field
                if not row.get('phone'):
                    logger.warning(f"Line {idx}: Missing phone number, skipping")
                    continue

                contacts.append(row)

        logger.info(f"✅ Read {len(contacts)} contacts from {file_path}")
        return contacts

    except FileNotFoundError:
        logger.error(f"❌ File not found: {file_path}")
        return []
    except Exception as e:
        logger.error(f"❌ Error reading file: {e}")
        return []

def read_excel_file(file_path: str) -> list:
    """
    Read contacts from Excel file.

    Requires: openpyxl or xlrd
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("❌ openpyxl not installed. Install with: pip install openpyxl")
        return []

    contacts = []
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Read data rows
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = dict(zip(headers, row))

            if not row_dict.get('phone'):
                logger.warning(f"Line {idx}: Missing phone number, skipping")
                continue

            contacts.append(row_dict)

        logger.info(f"✅ Read {len(contacts)} contacts from {file_path}")
        return contacts

    except Exception as e:
        logger.error(f"❌ Error reading Excel file: {e}")
        return []

def validate_contacts(contacts: list) -> tuple:
    """
    Validate contact data.

    Returns: (valid_contacts, invalid_count)
    """
    valid = []
    invalid_count = 0

    for contact_data in contacts:
        # Required field validation
        phone = contact_data.get('phone', '').strip()
        if not phone:
            invalid_count += 1
            continue

        # Basic phone format validation (remove spaces, check length)
        phone_cleaned = phone.replace(' ', '').replace('-', '').replace('.', '')
        if len(phone_cleaned) < 8:
            logger.warning(f"Invalid phone format: {phone}")
            invalid_count += 1
            continue

        valid.append(contact_data)

    logger.info(f"✅ Validated: {len(valid)} valid, {invalid_count} invalid")
    return valid, invalid_count

def insert_contacts_to_db(contacts_data: list) -> list:
    """
    Insert contacts into database.

    Returns: list of Contact objects with IDs
    """
    db = SessionLocal()
    contacts = []

    try:
        for data in contacts_data:
            contact = Contact(
                phone=data.get('phone', '').strip(),
                first_name=data.get('first_name', '').strip(),
                last_name=data.get('last_name', '').strip(),
                company=data.get('company', '').strip(),
                email=data.get('email', '').strip(),
                notes=data.get('notes', '').strip()
            )
            db.add(contact)
            contacts.append(contact)

        db.commit()

        # Refresh to get IDs
        for contact in contacts:
            db.refresh(contact)

        logger.info(f"✅ Inserted {len(contacts)} contacts into database")
        return contacts

    except Exception as e:
        logger.error(f"❌ Error inserting contacts: {e}")
        db.rollback()
        return []
    finally:
        db.close()

def main():
    parser = argparse.ArgumentParser(description="Import contacts from CSV/Excel")
    parser.add_argument("--source", required=True, help="Fichier CSV/Excel source")
    parser.add_argument("--campaign", help="Nom campagne (créer nouvelle ou ajouter)")
    parser.add_argument("--scenario", default="production", help="Scénario pour nouvelle campagne")

    args = parser.parse_args()

    logger.info(f"📥 Importing contacts from {args.source}...")

    # 1. Lire fichier (CSV ou Excel)
    file_ext = Path(args.source).suffix.lower()

    if file_ext == '.csv':
        contacts_data = read_csv_file(args.source)
    elif file_ext in ['.xlsx', '.xls']:
        contacts_data = read_excel_file(args.source)
    else:
        logger.error(f"❌ Unsupported file format: {file_ext}. Use .csv or .xlsx")
        sys.exit(1)

    if not contacts_data:
        logger.error("❌ No contacts read from file")
        sys.exit(1)

    # 2. Valider données
    valid_contacts, invalid_count = validate_contacts(contacts_data)

    if not valid_contacts:
        logger.error("❌ No valid contacts found")
        sys.exit(1)

    # 3. Insérer en DB
    contacts = insert_contacts_to_db(valid_contacts)

    if not contacts:
        logger.error("❌ Failed to insert contacts")
        sys.exit(1)

    # 4. Créer campagne si nécessaire
    if args.campaign:
        logger.info(f"📊 Creating campaign: {args.campaign}")

        contact_ids = [c.id for c in contacts]

        manager = CampaignManager()
        campaign_id = manager.create_campaign(
            name=args.campaign,
            contact_ids=contact_ids,
            scenario=args.scenario
        )

        logger.info(f"✅ Campaign created with ID: {campaign_id}")
        logger.info(f"   Launch with: python launch_campaign.py --campaign-id {campaign_id}")

    logger.info(f"\n✅ Import complete: {len(contacts)} contacts imported")

if __name__ == "__main__":
    main()
